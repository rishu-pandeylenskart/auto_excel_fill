
from __future__ import annotations

import datetime as dt
import re
import shutil
import tempfile
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

try:
    import win32com.client as win32
except Exception:
    win32 = None

SHEET_SHIPMENTS = "ShipmentDeatils"
SHEET_ITEMS = "ShipmentItemsDetails"

ALIASES = {
    "hawb": ["ARAMEX AWB", "AWB", "AWB ARAMEX", "HAWBNumber", "HAWB Number"],
    "reference": ["Reference Number", "Ref. Number", "Ref Number", "Order Reference",
                  "Order No", "Order Number", "OrderNo"],
    "contents": ["Contents", "Content", "Commodity", "Goods Description", "GoodsDescription"],
    "declared_value": ["Declared Value", "DeclaredValue", "Shipment Value", "Value"],
    "invoice": ["Invoice", "Invoice No", "Invoice Number", "InvoiceNo"],
    "qty": [
        "QTY", "Qty", "Quantity", "QTY.", "Qty.",
        "Item Qty", "Item Quantity", "Quantity Shipped", "No Of Pieces",
        "No. Of Pieces", "NoOfPieces"
    ],
    "currency": ["Currency", "Currency Code", "CurrencyCode"],
}

def norm_header(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = re.sub(r"[\s_\-]+", " ", s)
    s = re.sub(r"[^\w .]+", "", s)
    return s.strip()

def norm_text(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip()).casefold()

def display_text(v):
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""

def normalize_hawb(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s

def to_number(v):
    if v is None or str(v).strip() == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0

def parse_quantity(v):
    """Return quantity as a number, rejecting non-numeric nonblank values."""
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, bool):
        raise ValueError("Boolean is not a valid quantity")
    if isinstance(v, (int, float)):
        value = float(v)
    else:
        raw = str(v).strip().replace(",", "")
        try:
            value = float(raw)
        except Exception as e:
            raise ValueError(f"Invalid quantity value: {v!r}") from e
    if value < 0:
        raise ValueError(f"Quantity cannot be negative: {value}")
    return value

def find_column(headers, aliases):
    normalized = {norm_header(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        if norm_header(alias) in normalized:
            return normalized[norm_header(alias)]
    return None

def detect_columns(headers):
    return {k: find_column(headers, a) for k, a in ALIASES.items()}

def excel_check():
    if win32 is None:
        return False, "pywin32/win32com is unavailable in this portable runtime."
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        version = str(excel.Version)
        excel.Quit()
        return True, f"Microsoft Excel COM {version}"
    except Exception as e:
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        return False, str(e)

def excel_convert_to_xlsx(path: Path, temp_dir: Path) -> Path:
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        target = temp_dir / f"{path.stem}_source.xlsx"
        shutil.copy2(path, target)
        return target

    if win32 is None:
        raise RuntimeError("pywin32 is required for .xls/.xlsm source files.")

    target = temp_dir / f"{path.stem}_converted.xlsx"
    excel = wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True, UpdateLinks=0)
        wb.SaveAs(str(target.resolve()), FileFormat=51)  # .xlsx
        return target
    except Exception as e:
        raise RuntimeError(f"Excel could not convert '{path.name}' to working XLSX: {e}") from e
    finally:
        try:
            if wb:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass

def excel_save_as_xls(xlsx_path: Path, xls_path: Path):
    if win32 is None:
        raise RuntimeError("pywin32 is required to create legacy .xls output.")
    excel = wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(str(Path(xlsx_path).resolve()), ReadOnly=False, UpdateLinks=0)
        wb.SaveAs(str(Path(xls_path).resolve()), FileFormat=56)  # xlExcel8 / .xls
    except Exception as e:
        raise RuntimeError(f"Microsoft Excel could not create final .xls: {e}") from e
    finally:
        try:
            if wb:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass

def read_source(path: Path, temp_dir: Path):
    working = excel_convert_to_xlsx(path, temp_dir)
    wb = load_workbook(working, data_only=False)
    ws = None
    cols = None
    for candidate in wb.worksheets:
        headers = [c.value for c in candidate[1]]
        detected = detect_columns(headers)
        if detected["hawb"] is not None:
            ws, cols = candidate, detected
            break
    if ws is None:
        raise ValueError(f"{path.name}: no sheet containing an AWB/HAWB column was found.")

    missing = [k for k in ("hawb", "contents", "declared_value", "invoice") if cols[k] is None]
    if missing:
        raise ValueError(f"{path.name}: missing required source columns: {', '.join(missing)}")

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in vals):
            continue
        rows.append({
            "source_file": path.name,
            "source_row": r,
            "hawb": normalize_hawb(vals[cols["hawb"]]),
            "contents": display_text(vals[cols["contents"]]),
            "contents_norm": norm_text(vals[cols["contents"]]),
            "declared_value": to_number(vals[cols["declared_value"]]),
            "invoice": vals[cols["invoice"]],
            "reference": vals[cols["reference"]] if cols["reference"] is not None else None,
            "qty": parse_quantity(vals[cols["qty"]]) if cols["qty"] is not None else None,
            "currency": vals[cols["currency"]] if cols["currency"] is not None else None,
        })
    return rows

def resolve_hs(contents):
    n = norm_text(contents)
    if re.search(r"\bprescription[\s-]*lens\b", n, re.I):
        return "90015000"
    if re.search(r"\bpolarized\s+sunglasses\b", n, re.I):
        return "90041000"
    if re.search(r"\bprescription\s+eyeglasses\b", n, re.I):
        return "90049090"
    if n in ("eyeframe", "eyeframes"):
        return "90031100"
    return None

def template_col(ws, name):
    wanted = norm_header(name)
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(1, c).value) == wanted:
            return c
    return None

def template_cols(ws, names):
    return {name: template_col(ws, name) for name in names}

def capture_row2(ws):
    return [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

def clear_rows_after_2(ws):
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

def copy_complete_row2(ws, donor, target_row):
    """Master-template rule: row 2 values/formulas + formatting are filled down."""
    for c in range(1, ws.max_column + 1):
        src = ws.cell(2, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)

        value = donor[c - 1]
        if isinstance(value, str) and value.startswith("="):
            try:
                dst.value = Translator(value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = value
        else:
            dst.value = value

def write(ws, row, col, value):
    if col is not None:
        ws.cell(row, col).value = value

class ValidationFailure(Exception):
    def __init__(self, errors):
        self.errors = errors
        super().__init__(f"{len(errors)} validation error(s)")

def process(source_paths, template_path, output_dir, progress=lambda x: None):
    source_paths = [Path(x) for x in source_paths]
    template_path = Path(template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ok, detail = excel_check()
    if not ok:
        raise RuntimeError(
            "Microsoft Excel is required on the target PC. Excel COM test failed: " + detail
        )

    tmp = Path(tempfile.mkdtemp(prefix="shipment_portable_"))
    try:
        all_rows = []
        for i, src in enumerate(source_paths, 1):
            progress(f"Reading {i}/{len(source_paths)}: {src.name}")
            all_rows.extend(read_source(src, tmp))
        if not all_rows:
            raise ValueError("No data rows found in the selected source files.")

        shipments = defaultdict(list)
        for row in all_rows:
            if row["hawb"]:
                shipments[row["hawb"]].append(row)

        errors = []
        shipment_rows = []
        items_by_key = defaultdict(list)

        for hawb, rows in shipments.items():
            if not any(r["contents_norm"] for r in rows):
                errors.append({
                    "file": rows[0]["source_file"], "row": rows[0]["source_row"],
                    "hawb": hawb, "contents": "", "reason": "Contents is blank."
                })
                continue

            # ShipmentDeatils.Contents must overwrite the template row-2 value;
            # it is not a passive fill-down field. Use the first nonblank original
            # Contents value for the HAWB. All normalized contents are still grouped
            # separately for ShipmentItemsDetails.
            preferred_contents = next(
                (r["contents"] for r in rows if r["contents"]), None
            )
            shipment_rows.append({
                "hawb": hawb,
                "contents": preferred_contents,
                "reference": next((r["reference"] for r in rows if r["reference"] not in (None, "")), None),
                "invoice": next((r["invoice"] for r in rows if r["invoice"] not in (None, "")), None),
                "shipment_value": sum(r["declared_value"] for r in rows),
                "rows": rows,
            })
            for r in rows:
                if r["contents_norm"]:
                    items_by_key[(hawb, r["contents_norm"])].append(r)

        item_rows = []
        for (hawb, _), rows in items_by_key.items():
            contents = next((r["contents"] for r in rows if r["contents"]), "")
            hs = resolve_hs(contents)
            if hs is None:
                for r in rows:
                    errors.append({
                        "file": r["source_file"], "row": r["source_row"],
                        "hawb": hawb, "contents": r["contents"],
                        "reason": "Commodity has no defined HS-code rule."
                    })
                continue
            # Quantity rule:
            # - If a recognized quantity column exists in the source, SUM every
            #   numeric quantity for this HAWB + normalized Contents. Blank cells
            #   contribute 0; invalid text is rejected during source parsing.
            # - If no quantity column exists, COUNT the matching source rows.
            quantity_column_exists = any(
                r["qty"] is not None for r in rows
            )
            if quantity_column_exists:
                qty = sum((r["qty"] or 0) for r in rows)
            else:
                qty = len(rows)

            item_rows.append({
                "hawb": hawb, "contents": contents, "hs": hs,
                "qty": qty,
                "total_fob": sum(r["declared_value"] for r in rows),
            })

        if errors:
            raise ValidationFailure(errors)

        progress("Opening template...")
        template_xlsx = excel_convert_to_xlsx(template_path, tmp)
        wb = load_workbook(template_xlsx, data_only=False)

        if SHEET_SHIPMENTS not in wb.sheetnames or SHEET_ITEMS not in wb.sheetnames:
            raise ValueError(
                f"Template must contain {SHEET_SHIPMENTS} and {SHEET_ITEMS}. "
                f"Found: {', '.join(wb.sheetnames)}"
            )

        ws_s, ws_i = wb[SHEET_SHIPMENTS], wb[SHEET_ITEMS]
        s_cols = template_cols(ws_s, [
            "HAWBNumber", "Contents", "ShipmentValue", "ShipperReference1_OrderNO", "InvoiceNo", "InvoiceDate"
        ])
        i_cols = template_cols(ws_i, [
            "HAWBNumber", "CommodityType", "GoodsDescription", "IsMEIS", "HSCode",
            "Quantity", "InvoiceRatePerUnit", "TotalFOBValue", "TotalCESSPaid",
            "UnitOfMeasure", "ItemDetailsRef1", "ItemDetailsRef2", "ItemDetailsRef3",
            "SKU", "IGSTRate", "IGSTAmount"
        ])

        # CRITICAL: capture row 2 BEFORE clearing rows. Row 2 is never treated as blank.
        s_donor = capture_row2(ws_s)
        i_donor = capture_row2(ws_i)
        clear_rows_after_2(ws_s)
        clear_rows_after_2(ws_i)

        today = dt.date.today()

        progress("Filling ShipmentDeatils from row 2...")
        for row_no, rec in enumerate(shipment_rows, 2):
            copy_complete_row2(ws_s, s_donor, row_no)
            write(ws_s, row_no, s_cols.get("HAWBNumber"), rec["hawb"])
            write(ws_s, row_no, s_cols.get("Contents"), rec["contents"])
            write(ws_s, row_no, s_cols.get("ShipmentValue"), rec["shipment_value"])
            write(ws_s, row_no, s_cols.get("ShipperReference1_OrderNO"), rec["reference"])
            write(ws_s, row_no, s_cols.get("InvoiceNo"), rec["invoice"])
            write(ws_s, row_no, s_cols.get("InvoiceDate"), today)
            if s_cols.get("InvoiceDate"):
                ws_s.cell(row_no, s_cols["InvoiceDate"]).number_format = "dd-mm-yyyy"

        progress("Filling ShipmentItemsDetails from row 2...")
        for row_no, rec in enumerate(item_rows, 2):
            copy_complete_row2(ws_i, i_donor, row_no)
            write(ws_i, row_no, i_cols.get("HAWBNumber"), rec["hawb"])
            write(ws_i, row_no, i_cols.get("CommodityType"), rec["contents"])
            write(ws_i, row_no, i_cols.get("GoodsDescription"), rec["contents"])
            write(ws_i, row_no, i_cols.get("HSCode"), rec["hs"])
            write(ws_i, row_no, i_cols.get("Quantity"), rec["qty"])
            write(ws_i, row_no, i_cols.get("TotalFOBValue"), rec["total_fob"])

            # Preserve row-2 InvoiceRatePerUnit formula. Only fallback if row 2 did not have one.
            col = i_cols.get("InvoiceRatePerUnit")
            donor_formula = i_donor[col - 1] if col else None
            if col and not (isinstance(donor_formula, str) and donor_formula.startswith("=")):
                fob = i_cols.get("TotalFOBValue")
                qty = i_cols.get("Quantity")
                if fob and qty:
                    ws_i.cell(row_no, col).value = (
                        f'=IFERROR({get_column_letter(fob)}{row_no}/'
                        f'{get_column_letter(qty)}{row_no},"")'
                    )

        stamp = dt.date.today().strftime("%Y-%m-%d")
        working = tmp / f"Completed_Shipment_Template_{stamp}_working.xlsx"
        final_xls = output_dir / f"Completed_Shipment_Template_{stamp}.xls"

        progress("Saving working workbook...")
        wb.save(working)
        progress("Creating final legacy .xls with Microsoft Excel...")
        excel_save_as_xls(working, final_xls)

        progress("Done.")
        return final_xls

    finally:
        shutil.rmtree(tmp, ignore_errors=True)
